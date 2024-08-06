import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Carica il file Excel
file_path = 'DietApp.xlsx'
data = pd.read_excel(file_path)

# Funzione per cercare alimenti per nome
def search_food(food_name):
    return data[data['Nome'].str.contains(food_name, case=False)]

# Funzione per calcolare i macronutrienti totali
def calculate_totals(selected_foods):
    total_calories = sum(selected_foods['Calorie (kcal)'])
    total_proteins = sum(selected_foods['Proteine (g)'])
    total_carbs = sum(selected_foods['Carbo (g)'])
    total_fats = sum(selected_foods['Grassi (g)'])
    return total_calories, total_proteins, total_carbs, total_fats

# Funzione per calcolare i totali giornalieri
def calculate_daily_totals(meals):
    total_calories = 0
    total_proteins = 0
    total_carbs = 0
    total_fats = 0
    for meal, foods in meals.items():
        if foods:
            meal_data = pd.concat(foods)
            meal_totals = calculate_totals(meal_data)
            total_calories += meal_totals[0]
            total_proteins += meal_totals[1]
            total_carbs += meal_totals[2]
            total_fats += meal_totals[3]
    return total_calories, total_proteins, total_carbs, total_fats

# Funzione per aggiungere nuovi alimenti al file Excel
def add_new_food(name, calories, proteins, carbs, fats):
    new_data = pd.DataFrame({
        'Nome': [name],
        'Calorie (kcal)': [calories],
        'Proteine (g)': [proteins],
        'Carbo (g)': [carbs],
        'Grassi (g)': [fats]
    })
    global data
    data = pd.concat([data, new_data], ignore_index=True)
    data.to_excel(file_path, index=False)

# Funzione per combinare i dati dei pasti e dei totali giornalieri
def combine_meals_and_totals(meals, totals_df):
    all_meals_data = []
    for meal_name, foods in meals.items():
        if foods:
            meal_data = pd.concat(foods, ignore_index=True)
            meal_data['Meal'] = meal_name
            all_meals_data.append(meal_data)
    if all_meals_data:
        combined_meals_data = pd.concat(all_meals_data, ignore_index=True)
        combined_meals_data = combined_meals_data[['Meal', 'Nome', 'Calorie (kcal)', 'Proteine (g)', 'Carbo (g)', 'Grassi (g)', 'Quantity (g)']]
        
        # Aggiungi le colonne Nutrient, Total e Goal con valori None per mantenere la struttura
        combined_meals_data['Nutrient'] = None
        combined_meals_data['Total'] = None
        combined_meals_data['Goal'] = None

        # Aggiungi i totali giornalieri in fondo al DataFrame
        totals_rows = totals_df.shape[0]
        totals_df['Meal'] = [None] * totals_rows
        totals_df['Nome'] = [None] * totals_rows
        totals_df['Calorie (kcal)'] = [None] * totals_rows
        totals_df['Proteine (g)'] = [None] * totals_rows
        totals_df['Carbo (g)'] = [None] * totals_rows
        totals_df['Grassi (g)'] = [None] * totals_rows
        totals_df['Quantity (g)'] = [None] * totals_rows

        # Concatenare i DataFrame
        combined_data = pd.concat([combined_meals_data, totals_df], axis=0, ignore_index=True)
        
        # Riempire le celle vuote con valori appropriati
        combined_data.fillna('', inplace=True)
        
        return combined_data
    else:
        return pd.DataFrame(columns=['Meal', 'Nome', 'Calorie (kcal)', 'Proteine (g)', 'Carbo (g)', 'Grassi (g)', 'Quantity (g)', 'Nutrient', 'Total', 'Goal'])

# Funzione per salvare i pasti e i totali giornalieri in un unico foglio di Excel
def save_meals_and_totals_to_excel(sheet_name, meals, totals_df):
    combined_data = combine_meals_and_totals(meals, totals_df)
    with pd.ExcelWriter(file_path, mode='a', if_sheet_exists='replace') as writer:
        combined_data.to_excel(writer, sheet_name=sheet_name, index=False)

# Funzione per ottenere i nomi dei fogli di lavoro
def get_sheet_names(file_path):
    workbook = load_workbook(file_path, read_only=True)
    return workbook.sheetnames

# Funzione per caricare e visualizzare il contenuto di uno sheet
def load_sheet(sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df

# Interfaccia principale
st.title('Diet Tracker')

# Menu a tendina nella sidebar
option = st.sidebar.selectbox('Choose an option', ('Home', 'Add New Food', 'Set Daily Goals', 'View Saved Sheet'))

if option == 'Add New Food':
    # Sezione per aggiungere nuovi alimenti
    st.sidebar.header('Add New Food')
    with st.sidebar.form(key='add_food_form'):
        name = st.text_input('Food Name')
        calories = st.number_input('Calories (kcal) per 100g', min_value=0, step=1)
        proteins = st.number_input('Proteins (g) per 100g', min_value=0, step=1)
        carbs = st.number_input('Carbohydrates (g) per 100g', min_value=0, step=1)
        fats = st.number_input('Fats (g) per 100g', min_value=0, step=1)
        submit_button = st.form_submit_button(label='Add Food')

        if submit_button:
            add_new_food(name, calories, proteins, carbs, fats)
            st.sidebar.success(f'Added new food: {name}')

elif option == 'Set Daily Goals':
    # Input degli obiettivi giornalieri
    st.sidebar.header('Set Daily Goals')
    calories_goal = st.sidebar.number_input('Daily Calorie Goal (kcal)', min_value=0, max_value=10000, step=100)
    protein_goal = st.sidebar.number_input('Daily Protein Goal (g)', min_value=0, max_value=500, step=1)
    carbs_goal = st.sidebar.number_input('Daily Carbohydrate Goal (g)', min_value=0, max_value=1000, step=1)
    fat_goal = st.sidebar.number_input('Daily Fat Goal (g)', min_value=0, max_value=500, step=1)

    if st.sidebar.button('Save Goals'):
        st.session_state['goals'] = {
            'calories': calories_goal,
            'proteins': protein_goal,
            'carbs': carbs_goal,
            'fats': fat_goal
        }
        st.sidebar.success('Goals saved successfully!')

elif option == 'View Saved Sheet':
    # Sezione per visualizzare il contenuto di uno sheet salvato
    st.sidebar.header('View Saved Sheet')
    sheet_names = get_sheet_names(file_path)
    sheet_name = st.sidebar.selectbox('Select sheet name to view', sheet_names)
    if st.sidebar.button('Load Sheet'):
        if sheet_name:
            df = load_sheet(sheet_name)
            st.write(f'Contents of sheet: {sheet_name}')
            st.dataframe(df)
        else:
            st.sidebar.error('Please select a sheet name')

# Input del numero di pasti
st.sidebar.header('Set Number of Meals')
num_meals = st.sidebar.number_input('Number of Meals', min_value=3, max_value=10, step=1)

# Inizializza lo stato della sessione per i pasti e i nomi dei pasti
if 'meals' not in st.session_state:
    st.session_state['meals'] = {f'Meal {i}': [] for i in range(1, num_meals + 1)}
    st.session_state['meal_names'] = {f'Meal {i}': f'Meal {i}' for i in range(1, num_meals + 1)}
else:
    # Aggiunge nuovi pasti se il numero di pasti aumenta
    for i in range(1, num_meals + 1):
        if f'Meal {i}' not in st.session_state['meals']:
            st.session_state['meals'][f'Meal {i}'] = []
            st.session_state['meal_names'][f'Meal {i}'] = f'Meal {i}'

# Aggiunta di alimenti ai pasti
for meal in range(1, num_meals + 1):
    meal_key = f'Meal {meal}'
    meal_name = st.text_input(f'Enter name for {meal_key}', value=st.session_state['meal_names'][meal_key], key=f'name_{meal}')
    st.session_state['meal_names'][meal_key] = meal_name

    st.subheader(f'{meal_name}')
    food_name = st.text_input(f'Search food for {meal_name}', key=f'meal{meal}_food')
    if food_name:
        results = search_food(food_name)
        selected_food = st.selectbox(f'Select food for {meal_name}', results['Nome'], key=f'select_meal{meal}')
        if selected_food:
            quantity = st.number_input(f'Enter quantity (g) for {selected_food}', min_value=0, max_value=1000, step=1, key=f'quantity{meal}')
            if st.button(f'Add to {meal_name}', key=f'add_meal{meal}'):
                food_data = results[results['Nome'] == selected_food].copy()
                food_data['Calorie (kcal)'] = food_data['Calorie (kcal)'].astype(float) * quantity / 100
                food_data['Proteine (g)'] = food_data['Proteine (g)'].astype(float) * quantity / 100
                food_data['Carbo (g)'] = food_data['Carbo (g)'].astype(float) * quantity / 100
                food_data['Grassi (g)'] = food_data['Grassi (g)'].astype(float) * quantity / 100
                food_data['Quantity (g)'] = quantity
                st.session_state['meals'][meal_key].append(food_data)
                st.success(f'Added {selected_food} ({quantity}g) to {meal_name}')
                # Resetta i campi di input
                #st.query_params()

    # Mostra i dati del pasto corrente
    if st.session_state['meals'][meal_key]:
        st.write(f"Details for {meal_name}")
        meal_data = pd.concat(st.session_state['meals'][meal_key], ignore_index=True)
        for i, row in meal_data.iterrows():
            st.write(f"{row['Nome']} ({row['Quantity (g)']}g): {row['Calorie (kcal)']} kcal, {row['Proteine (g)']} g proteins, {row['Carbo (g)']} g carbs, {row['Grassi (g)']} g fats")
            if st.button(f'Remove item', key=f'remove_{meal}_{i}'):
                st.session_state['meals'][meal_key].pop(i)
                st.rerun()

        if st.button(f'Delete {meal_name}', key=f'delete_meal{meal}'):
            st.session_state['meals'][meal_key] = []
            st.success(f'{meal_name} deleted')
            st.rerun()

    # Aggiungi un separatore orizzontale
    st.markdown("---")

# Sezione per il conteggio totale
st.sidebar.header('Daily Totals')
total_calories, total_proteins, total_carbs, total_fats = calculate_daily_totals(st.session_state['meals'])

# Creazione del DataFrame per i totali giornalieri
totals_df = pd.DataFrame({
    'Nutrient': ['Calories', 'Proteins', 'Carbohydrates', 'Fats'],
    'Total': [total_calories, total_proteins, total_carbs, total_fats],
    'Goal': [
        st.session_state['goals']['calories'] if 'goals' in st.session_state else 'Not Set',
        st.session_state['goals']['proteins'] if 'goals' in st.session_state else 'Not Set',
        st.session_state['goals']['carbs'] if 'goals' in st.session_state else 'Not Set',
        st.session_state['goals']['fats'] if 'goals' in st.session_state else 'Not Set'
    ]
})

# Visualizzazione del DataFrame nella sidebar
st.sidebar.write(totals_df)

# Verifica degli obiettivi con barre di avanzamento
st.sidebar.header('Goal Progress')
if 'goals' in st.session_state:
    goals = st.session_state['goals']

    calories_progress = min(total_calories / goals['calories'], 1.0) if goals['calories'] > 0 else 0
    st.sidebar.progress(calories_progress)
    st.sidebar.write(f"Calories: {total_calories} / {goals['calories']} kcal")

    proteins_progress = min(total_proteins / goals['proteins'], 1.0) if goals['proteins'] > 0 else 0
    st.sidebar.progress(proteins_progress)
    st.sidebar.write(f"Proteins: {total_proteins} / {goals['proteins']} g")

    carbs_progress = min(total_carbs / goals['carbs'], 1.0) if goals['carbs'] > 0 else 0
    st.sidebar.progress(carbs_progress)
    st.sidebar.write(f"Carbohydrates: {total_carbs} / {goals['carbs']} g")

    fats_progress = min(total_fats / goals['fats'], 1.0) if goals['fats'] > 0 else 0
    st.sidebar.progress(fats_progress)
    st.sidebar.write(f"Fats: {total_fats} / {goals['fats']} g")

else:
    st.sidebar.warning('Please set your daily goals.')

# Sezione per salvare i pasti
st.header('Save Meals to Excel')
sheet_name = st.text_input('Enter sheet name')
if st.button('Save Meals'):
    if sheet_name:
        save_meals_and_totals_to_excel(sheet_name, st.session_state['meals'], totals_df)
        st.success(f'Meals saved to sheet: {sheet_name}')
    else:
        st.error('Please enter a sheet name')